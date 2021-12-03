import openpyxl


wb = openpyxl.load_workbook('./data.xlsx')
print(wb.sheetnames)
ws = wb['Sheet1']


print(ws['A3'].value)

# 按单元格拿数据（从左到右，从上到下）
print("\n按单元格拿数据（从左到右，从上到下）")
print(ws.values)    # 生成器
for row in ws.values:
   for value in row:
     print(value)

# 按行拿数据
print("\n按行拿数据（方法一）")
# for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=2, values_only=True):
for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=2):
    print(row)

# 按行拿数据
print("\n按行拿数据（方法二）")
for row in ws.rows:
    for col in row:
        print(col.value, end="\t")

    print()