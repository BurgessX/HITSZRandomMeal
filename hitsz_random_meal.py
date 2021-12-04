"""随机生成要吃什么"""
# 2021/10/26

import random
from typing import List
import openpyxl
import argparse


def random_stall(stalls:'list[Stall]'):
    """随机显示档口列表中的一个档口"""
    ind = random.randint(0, len(stalls)-1)
    print(f"{stalls[ind].cateen_name}{stalls[ind].floor}楼的{stalls[ind].name}。")


class Stall():
    """食堂档口"""
    def __init__(self, name, cateen_name:str, floor:int) -> None:
        self.name = name
        self.cateen_name = cateen_name
        self.floor = floor


class Cateen():
    """食堂"""
    def __init__(self, name, location, stalls:'list[Stall]'=None) -> None:
        """初始化"""
        self.name = name
        self.location = location

        if stalls is not None:
            self.stalls = stalls
            self.stall_num = len(stalls)
            self.stall_names = []
            for stall in stalls:
                self.stall_names.append(stall.name)
        else:
            self.stalls = []
            self.stall_num = 0
            self.stall_names = []

    def show_stalls(self):
        """显示食堂的所有档口"""
        print(f"{self.name}共{self.stall_num}个档口如下：")
        for stall_name in self.stall_names:
            print(f"- {stall_name}")

    def add_stall(self, stall:Stall):
        """新增档口"""
        if stall.cateen_name == self.name:
            self.stalls.append(stall)
            self.stall_num += 1
            self.stall_names.append(stall.name)
        else:
            print(f"新增档口失败！{stall.name}不属于{self.name}。")

    def delete_stall(self, stall_name):
        """删除档口"""
        sum = 0
        for stall in self.stalls:
            if stall.name == stall_name:
                self.stalls.remove(stall)
                self.stall_num -= 1
                self.stall_names.remove(stall_name)
                print(f"成功将{stall_name}从{self.name}中删除。")
                sum += 1

        if sum == 0:
            print(f"删除档口失败！{self.name}中没有{stall_name}。")
        else:
            print(f"共删除了{sum}个档口。")

    def random_stall(self):
        """随机显示食堂的一个档口"""
        ind = random.randint(0, self.stall_num - 1)
        print(self.stall_names[ind])




if __name__ == "__main__":

    ## Argparse脚本
    """例
    python hitsz_rondom_meal.py                         # 随机生成一个档口
    python hitsz_rondom_meal.py 荔园一食堂               # 随机生成一食堂中的一个档口
    python hitsz_rondom_meal.py 荔园一食堂 荔园二食堂     # 随机生成一食堂或二食堂的一个档口
    python hitsz_rondom_meal.py --show-all                      # 展示所有档口
    python hitsz_rondom_meal.py 荔园一食堂 --show-all            # 展示一食堂的所有档口
    """
    ap = argparse.ArgumentParser()
    ap.add_argument('cateens', nargs='*', default=['荔园一食堂', '荔园二食堂', '荔园三食堂', '荔园四食堂'], help='限定要在哪些食堂中随机，默认为所有食堂')
    ap.add_argument('-a', '--show-all', default=False, action='store_const', const=True, help='展示cateens的所有档口')
    args = ap.parse_args()
    # print(args.cateens)
    # print(args.show_all)


    ## 一、初始化
    # 导入表格
    wb = openpyxl.load_workbook('./data.xlsx')
    # print(wb.sheetnames)
    stall_ws = wb['档口表']

    # 初始化食堂
    hit_cateens = []
    hit_cateens.extend([
        Cateen('荔园一食堂', '哈工大'),
        Cateen('荔园二食堂', '哈工大'),
        Cateen('荔园三食堂', '哈工大'),
        Cateen('荔园四食堂', '哈工大')
        ])
    cateen_names = ['荔园一食堂', '荔园二食堂', '荔园三食堂', '荔园四食堂']

    # 载入档口数据
    for row in stall_ws.iter_rows(min_row=2):
        stall = Stall(row[0].value, row[1].value, row[2].value)
        if stall.cateen_name == '荔园一食堂':
            hit_cateens[0].add_stall(stall)
        elif stall.cateen_name == '荔园二食堂':
            hit_cateens[1].add_stall(stall)
        elif stall.cateen_name == '荔园三食堂':
            hit_cateens[2].add_stall(stall)
        elif stall.cateen_name == '荔园四食堂':
            hit_cateens[3].add_stall(stall)
        else:
            print('不属于哈工大食堂')
            continue


    ## 二、任务执行
    # 1.打印指定食堂的所有档口
    if args.show_all is True:
        for cateen in args.cateens:
            for i in range(len(cateen_names)):
                if hit_cateens[i].name == cateen:
                    hit_cateens[i].show_stalls()
                    break

    # 2.随机展示档口
    stalls_to_random = []   # 待随机的档口
    for cateen in args.cateens:
        for i in range(len(cateen_names)):
            if hit_cateens[i].name == cateen:
                stalls_to_random.extend(hit_cateens[i].stalls)
                break
    print("- 随机食堂列表：")
    print(args.cateens)
    print("- 随机档口列表：")
    print([stall.name for stall in stalls_to_random])
    print("- 随机生成的档口为：")
    random_stall(stalls_to_random)
